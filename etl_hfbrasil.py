"""
ETL Brasil — Preços Limão Tahiti (HF Brasil / Cepea — fonte oficial do Power BI)
================================================================================
Réplica fiel da query M "Base_Brasil" do Reporte Limão_finalversion.pbix:

  1. Baixa o xlsx de exportação do HF Brasil (mesma URL do Power Automate):
     produto=9, regiões 109-113 (Araraquara, Bebedouro, Limeira, Mogiana),
     periodicidade=diário, 2023 → ano corrente
  2. Filtra Produto = "Lima Ácida Tahiti - Colhida - Mercado"
  3. Preco_4_5kg = Preço / 6   (idêntico ao M — preço HF é R$/caixa 27,2kg)
  4. Semana = Date.WeekOfYear  (padrão Power Query: semana inicia DOMINGO,
     semana 1 = semana que contém 1º de janeiro — NÃO é semana ISO)
  5. tipo = "HF Brasil" (usado pelo frontend em drawBrasilHistorico)

Uso:
    python etl_hfbrasil.py [caminho_xlsx_local]   # arg opcional p/ teste offline

Saída:
    brasil_precos.csv  — preços diários por região
    Upsert → brasil_precos (Supabase, on_conflict=data,regiao,tipo)
"""

import csv
import io
import sys
from datetime import date, datetime, timedelta, timezone

import requests
import openpyxl

# ── CONFIG ─────────────────────────────────────────────────────────────────────
ANO_INICIAL = 2023
ANO_FINAL = date.today().year

URL = (
    "https://www.hfbrasil.org.br/br/estatistica/preco/exportar.aspx"
    "?produto=9"
    "&regiao[]=111&regiao[]=110&regiao[]=109&regiao[]=112&regiao[]=113"
    "&periodicidade=diario"
    f"&ano_inicial={ANO_INICIAL}&ano_final={ANO_FINAL}"
)

PRODUTO_FILTRO = "Lima Ácida Tahiti - Colhida - Mercado"
OUTPUT_CSV = "brasil_precos.csv"

HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/125.0.0.0 Safari/537.36"
    ),
    "Accept-Language": "pt-BR,pt;q=0.9",
}


# ── HELPERS ────────────────────────────────────────────────────────────────────
def week_of_year_pq(d: date) -> int:
    """
    Réplica de Date.WeekOfYear do Power Query (default):
    semana inicia no DOMINGO; semana 1 é a que contém 1º de janeiro.
    NÃO é semana ISO — mantido assim por fidelidade ao PBI.
    """
    jan1 = date(d.year, 1, 1)
    # domingo em (ou antes de) 1º jan — weekday(): Mon=0 ... Sun=6
    dias_desde_domingo = (jan1.weekday() + 1) % 7
    inicio_semana1 = jan1 - timedelta(days=dias_desde_domingo)
    return (d - inicio_semana1).days // 7 + 1


def baixar_xlsx() -> bytes:
    print(f"    GET {URL[:80]}...")
    r = requests.get(URL, headers=HEADERS, timeout=60)
    r.raise_for_status()
    ct = r.headers.get("content-type", "")
    if "spreadsheet" not in ct and not r.content[:2] == b"PK":
        raise RuntimeError(f"Resposta não é xlsx (content-type: {ct})")
    print(f"    OK — {len(r.content):,} bytes")
    return r.content


# ── PARSE ──────────────────────────────────────────────────────────────────────
def parse(xlsx_bytes: bytes) -> list[dict]:
    wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes), read_only=True)
    ws = wb["Sheet1"] if "Sheet1" in wb.sheetnames else wb[wb.sheetnames[0]]

    rows = ws.iter_rows(values_only=True)
    header = [str(c).strip() if c is not None else "" for c in next(rows)]
    idx = {name: i for i, name in enumerate(header)}
    for col in ("Produto", "Região", "Dia", "Mês", "Ano", "Preço"):
        if col not in idx:
            raise RuntimeError(f"Coluna '{col}' não encontrada. Header: {header}")

    extracted_at = datetime.now(timezone.utc).isoformat()
    records, unidades = [], set()

    for row in rows:
        if row is None or row[idx["Produto"]] is None:
            continue
        if str(row[idx["Produto"]]).strip() != PRODUTO_FILTRO:
            continue

        try:
            d = date(int(row[idx["Ano"]]), int(row[idx["Mês"]]), int(row[idx["Dia"]]))
            preco = float(row[idx["Preço"]])
        except (TypeError, ValueError):
            continue

        if "Unidade" in idx and row[idx["Unidade"]]:
            unidades.add(str(row[idx["Unidade"]]).strip())

        preco_4_5kg = preco / 6  # fiel ao M: [Preço] / 6

        records.append({
            "data":         d.isoformat(),
            "semana":       week_of_year_pq(d),     # Date.WeekOfYear (não-ISO)
            "ano":          d.year,
            "regiao":       str(row[idx["Região"]]).strip(),
            "tipo":         "HF Brasil",
            "preco_kg":     round(preco_4_5kg / 4.5, 4),
            "preco_4_5kg":  round(preco_4_5kg, 2),
            "extracted_at": extracted_at,
        })

    wb.close()

    # Dedup: o export do HF ocasionalmente repete linhas idênticas (mesma
    # data/região/preço). Mantém a primeira ocorrência por (data, regiao).
    vistos, dedup = set(), []
    for r in records:
        k = (r["data"], r["regiao"])
        if k not in vistos:
            vistos.add(k)
            dedup.append(r)
    if len(dedup) < len(records):
        print(f"    Dedup: {len(records) - len(dedup)} linhas repetidas removidas")
    records = dedup

    print(f"    Unidade(s) na fonte: {unidades or '—'}")
    print(f"    {len(records)} registros após filtro '{PRODUTO_FILTRO}'")
    return records


# ── MAIN ───────────────────────────────────────────────────────────────────────
def main():
    print("=" * 60)
    print("BRASIL ETL — HF Brasil (fonte oficial do Power BI)")
    print("=" * 60)

    if len(sys.argv) > 1:
        print(f"    Modo offline: {sys.argv[1]}")
        xlsx = open(sys.argv[1], "rb").read()
    else:
        xlsx = baixar_xlsx()

    records = parse(xlsx)
    if not records:
        print("Nenhum registro obtido — abortando sem tocar no banco.")
        sys.exit(1)

    # Salvar CSV
    with open(OUTPUT_CSV, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=records[0].keys())
        writer.writeheader()
        writer.writerows(records)
    print(f"    Salvo: {OUTPUT_CSV}")

    # Preview
    print("\nPreview (3 primeiros / 3 últimos):")
    for r in records[:3] + records[-3:]:
        print(f"  {r['data']} | S{r['semana']:02d} | {r['regiao']:25s} | "
              f"R$ {r['preco_4_5kg']:6.2f}/cx4,5kg")

    # Upsert Supabase
    try:
        from supabase_upsert import upsert
        result = upsert("brasil_precos", records, on_conflict="data,regiao,tipo")
        print(f"    Supabase: {result['inserted']} registros upserted")
        if result["errors"]:
            print(f"    Erros: {result['errors'][:2]}")
    except Exception as e:
        print(f"    Supabase skipped: {e}")

    print("\n" + "=" * 60)
    print("CONCLUIDO")


if __name__ == "__main__":
    main()
