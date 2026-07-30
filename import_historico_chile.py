"""
Importação histórica — Chile (ODEPA planilhas 2023/2024/2025)
==============================================================
Réplica fiel das queries M Chile2023/Chile2024/Chile2025 do PBI:

  - Lê 2023.xlsx / 2024.xlsx / 2025.xlsx (mesmos arquivos do ownCloud que o PBI usa)
  - Filtra Unidad de comercialización = "$/malla 18 kilos" (único filtro do PBI)
  - Mantém a granularidade completa: Mercado × Variedad × Calidad × Procedencia
    (o PBI faz Average por semana sobre TODAS essas linhas — pré-agregar muda a média)
  - semana / ano = colunas Semana / Ano do próprio xlsx (coincidem com ISO em 100%
    das linhas — validado 28/07/2026)
  - precio = Precio promedio / 18 → CLP/kg (mesma base do pipeline 2026;
    o frontend multiplica por 4.5/980 = fórmula Preço_4.5kg do PBI)
  - presentacion = "Variedad|Calidad|Procedencia" → mantém as linhas distintas
    na UNIQUE (fecha, mercado, presentacion)

Uso:
    python import_historico_chile.py

ATENÇÃO: roda DELETE (ano<=2025) + INSERT. Idempotente por substituição.
"""

import sys
from datetime import date, datetime, timezone
from pathlib import Path

import openpyxl
import requests

BASE_DIR = Path(__file__).parent
ARQUIVOS = ["2023.xlsx", "2024.xlsx", "2025.xlsx"]
UNIDAD_FILTRO = "$/malla 18 kilos"


def parse_fecha(v):
    if hasattr(v, "date"):
        return v.date()
    if isinstance(v, date):
        return v
    s = str(v).strip()[:10]
    for fmt in ("%Y-%m-%d", "%d-%m-%Y", "%d/%m/%Y"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            pass
    return None


def carregar(path: Path) -> list[dict]:
    wb = openpyxl.load_workbook(path, read_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    hdr = [str(c).strip() if c is not None else "" for c in rows[0]]
    def col(part):
        for i, h in enumerate(hdr):
            if part.lower() in h.lower():
                return i
        raise RuntimeError(f"coluna '{part}' não achada em {path.name}: {hdr}")

    i_sem, i_desde, i_ano  = col("Semana"), col("Desde"), col("Ano")
    i_merc, i_var, i_cal   = col("Mercado"), col("Variedad"), col("Calidad")
    i_proc, i_prec, i_uni  = col("Procedencia"), col("Precio promedio"), col("Unidad")

    extracted_at = datetime.now(timezone.utc).isoformat()
    out = []
    for r in rows[1:]:
        uni = str(r[i_uni]).strip() if r[i_uni] else ""
        if uni != UNIDAD_FILTRO:
            continue
        fecha = parse_fecha(r[i_desde])
        try:
            semana = int(r[i_sem])
            # PBI: Table.ReplaceValue(null → 2025) na coluna Ano do consolidado
            ano    = int(r[i_ano]) if r[i_ano] is not None else 2025
            precio = float(r[i_prec])
        except (TypeError, ValueError):
            continue
        if fecha is None or precio <= 0:
            continue

        pres = "|".join(str(r[i] or "").strip() for i in (i_var, i_cal, i_proc))
        out.append({
            "fecha":        fecha.isoformat(),
            "semana":       semana,
            "ano":          ano,
            "producto":     "LIMÓN",
            "mercado":      str(r[i_merc] or "").strip() or None,
            "presentacion": pres,
            "precio":       round(precio / 18, 2),   # CLP/kg (malla 18kg)
            "unidad":       "CLP/kg",
            "extracted_at": extracted_at,
        })
    return out


def main():
    records = []
    for nome in ARQUIVOS:
        path = BASE_DIR / nome
        if not path.exists():
            print(f"  ⚠ {nome} não encontrado — pulando")
            continue
        recs = carregar(path)
        print(f"  {nome}: {len(recs)} registros malla 18")
        records += recs

    if not records:
        print("Nada a importar."); sys.exit(1)

    # Dedup exato na chave (média se preços diferirem)
    from collections import defaultdict
    grupos = defaultdict(list)
    for r in records:
        grupos[(r["fecha"], r["mercado"], r["presentacion"])].append(r)
    dedup = []
    for regs in grupos.values():
        base = regs[0]
        if len(regs) > 1:
            base["precio"] = round(sum(x["precio"] for x in regs) / len(regs), 2)
        dedup.append(base)
    print(f"  Total após dedup: {len(dedup)} (de {len(records)})")

    # DELETE + INSERT
    env = dict(l.strip().split("=", 1) for l in open(BASE_DIR / ".env")
               if "=" in l and not l.startswith("#"))
    url, key = env["SUPABASE_URL"].strip(), env["SUPABASE_KEY"].strip()
    h = {"apikey": key, "Authorization": f"Bearer {key}"}
    r = requests.delete(f"{url}/rest/v1/chile_precos?ano=lte.2025", headers=h, timeout=60)
    print(f"  DELETE 2023-2025: {r.status_code}")

    from supabase_upsert import insert
    result = insert("chile_precos", dedup)
    print(f"  INSERT: {result['inserted']} registros")
    if result["errors"]:
        print(f"  Erros: {result['errors'][:2]}")


if __name__ == "__main__":
    main()
